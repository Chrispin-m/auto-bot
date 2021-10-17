import openpyxl
from selenium import webdriver
from pathlib import Path
from fpdf import FPDF
print("running...")
row_range=input("Enter the row range to generate report(eg 2-3):-->>")
row_range= row_range.split('-')
first_row = int(row_range[0])
last_row= int(row_range[1])

def getUrls():
    excel_file=Path('excels', "Script_23514.xlsx")
    workbook=openpyxl.load_workbook(excel_file)
    sheet=workbook.active
    #target column
    tcol = sheet['AK']
    ilist = list(range(first_row - 1, last_row))
    urls=[]
    for i in ilist:
        urls.append(tcol[i].value)

    return urls
def sNames():
     excel_file = Path('excels', "Script_23514.xlsx")
     workbook = openpyxl.load_workbook(excel_file)
     sheet = workbook.active
     idcol = sheet['A']  # id Column
     ilist = list(range(first_row - 1, last_row))
     sname=[]
     for i in ilist:
         sname.append(idcol[i].value)

     return sname


def getNatch():
    excel_file = Path('excels', "Script_23514.xlsx")
    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook.active
    ncol = sheet['D']  # id Column
    ilist = list(range(first_row - 1, last_row))

    natchnames = []  # sccreenshot names
    for i in ilist:
        natchnames.append(ncol[i].value)

    return natchnames
def getArnede1():
    excel_file = Path('excels', "Script_23514.xlsx")
    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook.active
    ncol = sheet['AJ']  #  Column
    ilist = list(range(first_row - 1, last_row))
    a1 = []
    for i in ilist:
        a1.append(ncol[i].value)

    return a1

def getArnede2():
    excel_file = Path('excels', "Script_23514.xlsx")
    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook.active
    ncol = sheet['AL']  #  Column
    ilist = list(range(first_row - 1, last_row))
    a2 = []
    for i in ilist:
        a2.append(ncol[i].value)
    return a2

def getArnede3():
    excel_file = Path('excels', "Script_23514.xlsx")
    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook.active
    ncol = sheet['AM']  #  Column
    ilist = list(range(first_row - 1, last_row))
    a3 = []
    for i in ilist:
        a3.append(ncol[i].value)
    return a3
print("Getting Anrede 1-3...")
ar1= getArnede1()
ar1.remove(ar1[0])
ar2= getArnede2()
ar2.remove(ar2[0])
ar3= getArnede3()
ar3.remove(ar3[0])

#print(ar1)
#print(ar2)
#print(ar3)


nnames=getNatch()
nnames.remove(nnames[0])
urls = getUrls()
scnames = sNames()
print(urls)
#print(scnames)
from PIL import Image #pip install Pillow
driver = webdriver.Chrome(executable_path = "chromedriver.exe")
urlheader = "https://www."
def Screenshot(url,sname): #sname = screenshot name/name of the screenshot file
    driver.get(urlheader+url)
    driver.save_screenshot("{}.png".format(sname))
#Screenshot(url,"test")
#obtain the screenshot of each url
urls.remove(urls[0])
scnames.remove(scnames[0])
for i in range(len(urls)):
    print("Saving Screenshot {}".format(scnames[i]))
    try:
        Screenshot(urls[i],scnames[i])
    except:
        print("{} failed to load".format(urls[i]))

#------------------------------------------------------------------------ Screenshot saved

for j in range(len(urls)):
    def getTextData():
        excel_file = Path('excels', "DMC template.xlsx")
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook.active
        textdata = []
        second_row = list(sheet.rows)[1]
        for i in second_row:
            textdata.append(i.value)
        return textdata


    '''
    def createPdf(pdfname):
        pdf = FPDF('P', 'mm', 'Letter')
        pdf.set_auto_page_break(auto=True, margin=15)
        pdf.add_page()
        pdf.set_font('helvetica', 'BIU', 16)
        title = str(getTextData()[0].encode('UTF-8'))
        pdf.cell(0, 10, f'{title}', ln=True)
        pdf.output(pdfname)
    '''

    # createPdf('pdfr.pdf')

    title = ' '


    class PDF(FPDF):
        def header(self):
            # font
            self.set_font('helvetica', 'B', 15)
            # Calculate width of title and position
            title_w = self.get_string_width(title) + 6
            doc_w = self.w
            self.set_x((doc_w - title_w) / 2)
            # colors of frame, background, and text
            self.set_draw_color(255, 255, 255)  # border = white
            # self.set_fill_color(230, 230, 0)  # background = yellow
            self.set_text_color(220, 50, 50)  # text = red
            # Thickness of frame (border)
            self.set_line_width(1)
            # Title
            self.cell(title_w, 10, title, border=1, ln=1, align='C')
            # Line break
            self.ln(10)

        # Page footer
        def footer(self):
            # Set position of the footer
            self.set_y(-15)
            # set font
            self.set_font('Arial', 'I', 8)
            # Set font color grey
            self.set_text_color(169, 169, 169)
            # Page number
            self.cell(0, 10, f'Page {self.page_no()}', align='C')

        # Adding chapter title to start of each chapter
        def Title(self, title):
            # set font
            self.set_font('Arial', 'B', 26)
            # background color

            self.cell(0, 5, title, ln=1)
            # line break
            self.ln()

        def pdf_bodyy(self, txt):
            # set font
            self.set_font('Arial', '', 19)
            # insert text
            if txt.endswith("fehlt"):
                text = "kritisches Element fehlt"  # part of cell context to put in bold, with padding matching the word horizontal position
                x,y = pdf.x, pdf.y

                pdf.set_xy(x, y)  # positioning FPDF to re-draw the same cell
                pdf.set_font(style="B")  # switching to bold
                pdf.multi_cell(0, 5, text)
                pdf.set_font(style="")  # switching back to regular
            else:
                self.multi_cell(0, 5, txt)
            pdf.cell(0, 0, '', ln=True)
            # line break
            self.ln()

        # Chapter content
        def pdf_body(self, txt):
            # set font
            self.set_font('Arial', '', 11)
            # insert text
            self.multi_cell(0, 5, txt)
            pdf.cell(0, 0, '', ln=True)
            # line break
            self.ln()
            # self.set_font('times', 'I', 12)
            # self.cell(0, 5, ' ')

        def print_text(self, txt):
            self.pdf_body(txt)

        def print_textt(self, txt):
            self.pdf_bodyy(txt)

        def add_image(self, path):
            pdf.image(path, x=50, y=50, w=100)
            pdf.ln(70)


    # Create a PDF object
    pdf = PDF('P', 'mm', 'Letter')

    # get total page numbers
    pdf.alias_nb_pages()

    # Set auto page break
    pdf.set_auto_page_break(auto=True, margin=15)

    # Add Page
    pdf.add_page()
    '''
    try:
        text = str(getTextData()[1])
    except:
        text = str(getTextData()[1].encode('UTF-8'))
    '''
    # pdf.Title("hi") my speed is slow

    # title  to remobe the b and '
    textdata = getTextData()
    t = str(textdata[0].encode('ascii', 'ignore')).replace("Nachname", nnames[j]).replace("Anrede2", ar2[j])
    pdf.Title(t[1:len(t) - 1].replace("'", "") + "...")
    pdf.print_textt("..."+(str(textdata[1]).replace("'", "")[1:len(t) - 1]).replace("Nachname", nnames[j]).replace("\\", ""))
    imagename = f'{scnames[j]}.png'
    print("Loading Image to pdf..")
    try:
        pdf.add_image(imagename)
    except:
        print(f'{imagename}.png does not exist ------ the url to obtain the screenshot returned an error')
        pass

#
    # print(tex)
    #print(len(textdata) - 1)
    for i in range(2, len(textdata)):
        #print(textdata[i])
        tex = str(textdata[i]).encode('ascii', 'ignore') #.replace("\\xe4","ä") #german character encoding
        pdf.print_text((str(tex).replace("'", "")[1:len(tex)]).replace("Nachname",nnames[j]).replace("\\", "").replace("Anrede 1", ar1[j]).replace("Anrede 3", ar3[j])) #.replace("\\xfc", "ü").replace("\\xf6", "ö").replace("\\xdf", "ß"))
        pdf.print_text('\n')

        #print("+++++" + str(tex).replace("'", "")[1:len(tex)])

    pdf.output('{}.pdf'.format(scnames[j]))


driver.close() # should close the chrome window after execution.


